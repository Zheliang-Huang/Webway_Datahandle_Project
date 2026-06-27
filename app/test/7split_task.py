import json
import os
import platform

import xlwt


def traverse_dir(file_path, path_read, ext):
    # 该方法可以递归遍历路径下的子文件，可指定格式

    # file_path: 待遍历路径
    # path_read: 子文件列表
    # ext：格式为列表，可写入指定格式后缀

    # 例：
    # src = r'D:\黄哲良\Python-dataHandle\爬数据\刘思汐_语料清洗（MNBVC）20230706\dst'
    # dst = r'D:\黄哲良\Python-dataHandle\爬数据\刘思汐_语料清洗（MNBVC）20230706\txt'
    # json_paths =[]
    # traverse_dir(src,json_paths,['.jsonl'])

    if platform.system() == "Windows":
        file_sep = '\\'
    elif platform.system() == "Linux":
        file_sep = '/'
    # put file name from file_path in temp_list
    temp_list = os.listdir(file_path)
    for temp_list_nor in temp_list:
        if os.path.isfile(file_path + file_sep + temp_list_nor):
            temp_path = file_path + file_sep + temp_list_nor
            if len(ext) == 0:
                path_read.append(temp_path)
            else:
                ext_dic = {key: 1 for key in ext}
                if os.path.splitext(temp_path)[-1] in ext_dic:
                    path_read.append(temp_path)
                else:
                    continue
        else:
            traverse_dir(file_path + file_sep + temp_list_nor,
                         path_read, ext)  # loop traversal


def create_excel(heading, content_list, savepath):
    import xlsxwriter
    import xlwt
    from tqdm import tqdm
    if len(content_list) < 65536:
        wbook = xlwt.Workbook(encoding='utf-8', style_compression=0)
        wsheet = wbook.add_sheet('sheet1', cell_overwrite_ok=True)
        for i in range(0, len(heading)):
            wsheet.write(0, i, heading[i])
        for row in tqdm(range(0, len(content_list)), desc='Creating Excel', unit='row', miniters=1, bar_format='{l_bar}{bar:20}{r_bar}{bar:-10b}'):
            contents = content_list[row]
            for idx in range(0, len(contents)):
                wsheet.write(row+1, idx, contents[idx])
        wbook.save(savepath.replace('.xlsx', '.xls'))
    else:
        wbook = xlsxwriter.Workbook(savepath.replace('.xls', '.xlsx'))
        wsheet = wbook.add_worksheet('sheet1')
        for i in range(0, len(heading)):
            wsheet.write(0, i, heading[i])
        for row in tqdm(range(0, len(content_list)), desc='Creating Excel', unit='row', miniters=1, bar_format='{l_bar}{bar:20}{r_bar}{bar:-10b}'):
            contents = content_list[row]
            for idx in range(0, len(contents)):
                wsheet.write(row+1, idx, contents[idx])
        wbook.close()


def parse_OCR_json(json_format, red_list):
    unique_values = set()
    char_location_list = []
    # 正文解析
    if len(json_format["referenceSignInstances"]["正文"]) != 0:
        main = json_format['referenceSignInstances']["正文"]
        if type(main) is dict:
            for line in main['original_position']:
                # # 正文成分标签与坐标
                # component_word = '正文'
                # component_points = line.get('points')
                # component_location = str(int(component_points[0][0]))+ ' ' +str(int(component_points[0][1]))+ ' ' +str(int(component_points[2][0]))+ ' ' +str(int(component_points[2][1]))
                # component = component_location + ' ' + component_word
                # if component not in unique_values:
                #     char_location_list.append(component)
                #     unique_values.add(component)
                for charline in line['charlines']:
                    # 获取成分
                    component_word = '正文'
                    component_points = charline.get('points')
                    component_location = str(int(component_points[0][0])) + ' ' + str(int(component_points[0][1])) + ' ' + str(
                        int(component_points[2][0])) + ' ' + str(int(component_points[2][1]))
                    component = component_location + ' ' + component_word
                    if component not in unique_values:
                        # char_location_list.append(component)
                        unique_values.add(component)
                    # 判断大小字
                    if 'font_size' in charline and charline['font_size'] == 'large':
                        for text in charline['texts']:
                            word = text['word']
                            bbox = text['char_bbox']
                            confident = (text.get('topk').get(word))*100
                            if confident < 99:
                                red_list.append(word)
                            char = str(int(bbox[0][0])) + ' ' + str(int(bbox[0][1])) + ' ' + str(
                                int(bbox[2][0])) + ' ' + str(int(bbox[2][1])) + ' ' + word
                            if char not in unique_values:
                                char_location_list.append(char)
                                unique_values.add(char)
                    if 'font_size' in charline and charline['font_size'] == 'small':
                        for text in charline['texts']['small_print']:
                            word = text['word']
                            bbox = text['char_bbox']
                            confident = (text.get('topk').get(word))*100
                            if confident < 99:
                                red_list.append(word)
                            char = str(int(bbox[0][0])) + ' ' + str(int(bbox[0][1])) + ' ' + str(
                                int(bbox[2][0])) + ' ' + str(int(bbox[2][1])) + ' ' + word
                            if char not in unique_values:
                                char_location_list.append(char)
                                unique_values.add(char)
                    if 'font_size' in charline and charline['font_size'] == 'media':
                        for text in charline['texts']:
                            word = text['word']
                            bbox = text['char_bbox']
                            confident = (text.get('topk').get(word))*100
                            if confident < 99:
                                red_list.append(word)
                            char = str(int(bbox[0][0])) + ' ' + str(int(bbox[0][1])) + ' ' + str(
                                int(bbox[2][0])) + ' ' + str(int(bbox[2][1])) + ' ' + word
                            if char not in unique_values:
                                char_location_list.append(char)
                                unique_values.add(char)
        elif type(main) is list:
            for main_dict in main:
                if 'original_position' in main_dict:
                    for line in main_dict['original_position']:
                        # 正文成分标签与坐标
                        # component_word = '正文'
                        # component_points = line.get('points')
                        # component_location = str(int(component_points[0][0]))+ ' ' +str(int(component_points[0][1]))+ ' ' +str(int(component_points[2][0]))+ ' ' +str(int(component_points[2][1]))
                        # component = component_location + ' ' + component_word
                        # if component not in unique_values:
                        #     char_location_list.append(component)
                        #     unique_values.add(component)
                        for charline in line['charlines']:
                            # 获取成分
                            component_word = '正文'
                            component_points = charline.get('points')
                            component_location = str(int(component_points[0][0])) + ' ' + str(int(component_points[0][1])) + ' ' + str(
                                int(component_points[2][0])) + ' ' + str(int(component_points[2][1]))
                            component = component_location + ' ' + component_word
                            if component not in unique_values:
                                # char_location_list.append(component)
                                unique_values.add(component)
                            # 判断大小字
                            if 'font_size' in charline and charline['font_size'] == 'large':
                                for text in charline['texts']:
                                    word = text['word']
                                    bbox = text['char_bbox']
                                    confident = (
                                        text.get('topk').get(word))*100
                                    if confident < 99:
                                        red_list.append(word)
                                    char = str(int(bbox[0][0])) + ' ' + str(int(bbox[0][1])) + ' ' + str(
                                        int(bbox[2][0])) + ' ' + str(int(bbox[2][1])) + ' ' + word
                                    if char not in unique_values:
                                        char_location_list.append(char)
                                        unique_values.add(char)
                            if 'font_size' in charline and charline['font_size'] == 'small':
                                for text in charline['texts']['small_print']:
                                    word = text['word']
                                    bbox = text['char_bbox']
                                    confident = (
                                        text.get('topk').get(word))*100
                                    if confident < 99:
                                        red_list.append(word)
                                    char = str(int(bbox[0][0])) + ' ' + str(int(bbox[0][1])) + ' ' + str(
                                        int(bbox[2][0])) + ' ' + str(int(bbox[2][1])) + ' ' + word
                                    if char not in unique_values:
                                        char_location_list.append(char)
                                        unique_values.add(char)
    # 表格解析
    if len(json_format["referenceSignInstances"]["表格"]) != 0:
        for row in json_format["referenceSignInstances"]["表格"]:
            for cell in row['blocks']:
                # 表格成分标签与坐标
                cell_name = '表格正文'
                cell_location = str(int(cell['xmin'])) + ' ' + str(int(cell['ymin'])) + ' ' + str(
                    int(cell['xmax'])) + ' ' + str(int(cell['ymax'])) + ' ' + cell_name
                if cell_location not in unique_values:
                    # char_location_list.append(cell_location)
                    unique_values.add(cell_location)
                if 'original_position' in cell and len(cell['original_position']) != 0:
                    for original_position in cell['original_position']:
                        for charline in original_position['charlines']:
                            # 判断大小字
                            if 'font_size' in charline and charline['font_size'] == 'large':
                                for text in charline['texts']:
                                    word = text['word']
                                    bbox = text['char_bbox']
                                    confident = (
                                        text.get('topk').get(word))*100
                                    if confident < 99:
                                        red_list.append(word)
                                    char = str(int(bbox[0][0])) + ' ' + str(int(bbox[0][1])) + ' ' + str(
                                        int(bbox[2][0])) + ' ' + str(int(bbox[2][1])) + ' ' + word
                                    if char not in unique_values:
                                        char_location_list.append(char)
                                        unique_values.add(char)
                            if 'font_size' in charline and charline['font_size'] == 'small':
                                for text in charline['texts']['small_print']:
                                    word = text['word']
                                    bbox = text['char_bbox']
                                    confident = (
                                        text.get('topk').get(word))*100
                                    if confident < 99:
                                        red_list.append(word)
                                    char = str(int(bbox[0][0])) + ' ' + str(int(bbox[0][1])) + ' ' + str(
                                        int(bbox[2][0])) + ' ' + str(int(bbox[2][1])) + ' ' + word
                                    if char not in unique_values:
                                        char_location_list.append(char)
                                        unique_values.add(char)
    # 天头解析
    if len(json_format["referenceSignInstances"]["天头"]) != 0:
        main = json_format['referenceSignInstances']["天头"]
        if type(main) is dict:
            for line in main['original_position']:
                for charline in line['charlines']:
                    # 获取成分
                    component_word = '天头'
                    component_points = charline.get('points')
                    component_location = str(int(component_points[0][0])) + ' ' + str(int(component_points[0][1])) + ' ' + str(
                        int(component_points[2][0])) + ' ' + str(int(component_points[2][1]))
                    component = component_location + ' ' + component_word
                    if component not in unique_values:
                        # char_location_list.append(component)
                        unique_values.add(component)
                    # 判断大小字
                    if 'font_size' in charline and charline['font_size'] == 'large':
                        for text in charline['texts']:
                            word = text['word']
                            bbox = text['char_bbox']
                            confident = (text.get('topk').get(word))*100
                            if confident < 99:
                                red_list.append(word)
                            char = str(int(bbox[0][0])) + ' ' + str(int(bbox[0][1])) + ' ' + str(
                                int(bbox[2][0])) + ' ' + str(int(bbox[2][1])) + ' ' + word
                            if char not in unique_values:
                                char_location_list.append(char)
                                unique_values.add(char)
                    if 'font_size' in charline and charline['font_size'] == 'small':
                        for text in charline['texts']['small_print']:
                            word = text['word']
                            bbox = text['char_bbox']
                            confident = (text.get('topk').get(word))*100
                            if confident < 99:
                                red_list.append(word)
                            char = str(int(bbox[0][0])) + ' ' + str(int(bbox[0][1])) + ' ' + str(
                                int(bbox[2][0])) + ' ' + str(int(bbox[2][1])) + ' ' + word
                            if char not in unique_values:
                                char_location_list.append(char)
                                unique_values.add(char)
    if '版心' in json_format["layoutAnalysis"]:
        for banxin in json_format["layoutAnalysis"]["版心"]:
            component_word = '版心'
            component_location = str(int(banxin[0])) + ' ' + str(
                int(banxin[1])) + ' ' + str(int(banxin[2])) + ' ' + str(int(banxin[3]))
            component = component_location + ' ' + component_word
            if component not in unique_values:
                # char_location_list.append(component)
                unique_values.add(component)
    return char_location_list


def readTxtFile(filepath):
    # 不要把open放在try中，以防止打开失败，那么就不用关闭了
    file_object = open(filepath, "r", encoding='UTF-8', errors='ignore')
    try:
        # file_context是一个string，读取完后，就失去了对test.txt的文件引用
        file_context = file_object.readlines()
        xmlstr = ""
        for lineStr in file_context:
            # datalist.append(lineStr)
            xmlstr = xmlstr + lineStr
        return xmlstr
    finally:
        file_object.close()


def read_excel(filePath, col_num=None, col_num_2=None):
    import os

    import xlrd
    from openpyxl import load_workbook
    ext = os.path.splitext(filePath)[-1]
    if ext == '.xlsx':
        wb = load_workbook(filePath)
        sheets = wb.worksheets   # 获取当前所有的sheet
        sheet = sheets[0]
        if col_num is None:
            output = []
            res_list = []
            for row in sheet.iter_rows(min_row=2):
                for col in range(0, sheet.max_column):
                    target = row[col].value
                    res_list.append(target)
                output.append(res_list)
                res_list = []
            return output
        if col_num_2 is None:
            res_list = []
            for row in sheet.iter_rows(min_row=2):

                target = row[col_num].value
                res_list.append(target)
            return res_list
        else:
            res_dic = {}
            for row in sheet.iter_rows(min_row=2):
                key = row[col_num].value
                target = row[col_num_2].value
                res_dic[key] = target
            return res_dic
    elif ext == '.xls':
        book = xlrd.open_workbook(filePath)
        book.sheets()
        sheet = book.sheet_by_index(0)
        # sheet.ncols
        if col_num is None:
            output = []
            res_list = []
            for row in range(1, sheet.nrows):
                for col in range(0, sheet.ncols):
                    target = sheet.cell_value(row, col)
                    res_list.append(target)
                output.append(res_list)
                res_list = []
            return output
        if col_num_2 is None:
            res_list = []
            for row in range(1, sheet.nrows):
                target = sheet.cell_value(row, col_num)
                res_list.append(target)
            return res_list
        else:
            res_dic = {}
            for row in range(1, sheet.nrows):
                key = sheet.cell_value(row, col_num)
                target = sheet.cell_value(row, col_num_2)
                res_dic[key] = target
            return res_dic


source_list = [
    r"F:\大数据中心\33 - 明代别集数据库\5-待OCR识别+统计字数\20260420_to"
    ]
excel_path = r'F:\Python37\py_file\图片数量&字数表20260420.xls'
zsk_list = read_excel(excel_path, col_num=1)
output = r'F:\Python37\py_file\拆分任务20260420.xls'
excel_list = []
for source in source_list:
    print(source)

    total_num = 0
    print('开始遍历子文件夹')

    num = 1
    for each_book in os.listdir(source):
        zsk_num = each_book.split(' ')[0]
        if zsk_num not in zsk_list:
            print('不在本次任务清单内')
            num += 1
            continue
        print('已完成：{}，共{}'.format(num, len(os.listdir(source))))
        full_path_each = os.path.join(source, each_book)
        if os.path.isdir(full_path_each):
            json_list = []
            traverse_dir(full_path_each, json_list, ['.json'])
            char_num = 0
            red_num = 0
            json_task_list = []
            for json_path in json_list:
                try:
                    json_read = readTxtFile(json_path)
                    json_get = json.loads(json_read)
                    if json_get is not None:
                        json_content = json_get['Images']
                        red_list = []
                        char_list = parse_OCR_json(json_content, red_list)
                        red_num += len(red_list)
                        char_num += len(char_list)
                        if char_num >= 50000:
                            json_task_list.append(os.path.basename(json_path))
                            excel_list.append([zsk_num, ','.join(json_task_list), char_num, red_num])
                            json_task_list = []
                            char_num = 0
                            red_num = 0
                        else:
                            json_task_list.append(os.path.basename(json_path))
                except:
                    continue
            if len(json_task_list) != 0:
                excel_list.append(
                    [zsk_num, ','.join(json_task_list), char_num, red_num])
            num += 1
create_excel(['任务标题', '处理范围', '任务量(千字)', '标红数量'], excel_list, output)
