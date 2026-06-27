from typing import List,Dict


def excel_col_to_number(col_letter: str) -> int:
    """将 Excel 列字母转为数字，如 'A'->1, 'Z'->26, 'AA'->27"""
    num = 0
    for ch in col_letter.upper():  # 确保大写
        if not 'A' <= ch <= 'Z':
            raise ValueError(f"无效列字母: {col_letter}")
        num = num * 26 + (ord(ch) - ord('A') + 1)
    return num


def read_excel(filePath, col_num=None, col_num_2=None) -> [list, dict]:
    import xlrd
    import os
    from openpyxl import load_workbook
    ext = os.path.splitext(filePath)[-1]
    if ext == '.xlsx':
        wb = load_workbook(filePath)
        sheets = wb.worksheets   # 获取当前所有的sheet
        sheet = sheets[0]
        if col_num is None:
            output = []
            res_list = []
            for row in sheet.iter_rows(min_row=2):
                for col in range(0, sheet.max_column):
                    target = row[col].value
                    res_list.append(target)
                output.append(res_list)
                res_list = []
            return output
        if col_num_2 is None:
            res_list = []
            for row in sheet.iter_rows(min_row=2):

                target = row[col_num].value
                res_list.append(target)
            return res_list
        else:
            res_dic = {}
            for row in sheet.iter_rows(min_row=2):
                key = row[col_num].value
                target = row[col_num_2].value
                res_dic[key] = target
            return res_dic
    elif ext == '.xls':
        book = xlrd.open_workbook(filePath)
        book.sheets()
        sheet = book.sheet_by_index(0)
        # sheet.ncols
        if col_num is None:
            output = []
            res_list = []
            for row in range(1, sheet.nrows):
                for col in range(0, sheet.ncols):
                    target = sheet.cell_value(row, col)
                    res_list.append(target)
                output.append(res_list)
                res_list = []
            return output
        if col_num_2 is None:
            res_list = []
            for row in range(1, sheet.nrows):
                target = sheet.cell_value(row, col_num)
                res_list.append(target)
            return res_list
        else:
            res_dic = {}
            for row in range(1, sheet.nrows):
                key = sheet.cell_value(row, col_num)
                target = sheet.cell_value(row, col_num_2)
                res_dic[key] = target
            return res_dic


def create_excel(heading, content_list, savepath):
    import os
    import xlsxwriter
    import xlwt
    from tqdm import tqdm
    wbook = xlsxwriter.Workbook(os.path.splitext(savepath)[0]+'.xlsx')
    wsheet = wbook.add_worksheet('sheet1')
    for i in range(0, len(heading)):
        wsheet.write(0, i, heading[i])
    for row in tqdm(range(0, len(content_list)), desc='Creating Excel', unit='row', miniters=1, bar_format='{l_bar}{bar:20}{r_bar}{bar:-10b}'):
        contents = content_list[row]
        for idx in range(0, len(contents)):
            wsheet.write(row+1, idx, contents[idx])
    wbook.close()


def create_excel_sheets(heading, content_dic, savepath):
    import os
    import xlsxwriter
    import xlwt
    from tqdm import tqdm
    wbook = xlsxwriter.Workbook(os.path.splitext(savepath)[0]+'.xlsx')
    for idx, key in enumerate(content_dic):
        content_list = content_dic[key]
        wsheet = wbook.add_worksheet(key)
        for i in range(0, len(heading[idx])):
            wsheet.write(0, i, heading[idx][i])
        for row in tqdm(range(0, len(content_list)), desc='Creating Excel', unit='row', miniters=1, bar_format='{l_bar}{bar:20}{r_bar}{bar:-10b}'):
            contents = content_list[row]
            for idx in range(0, len(contents)):
                wsheet.write(row+1, idx, contents[idx])
    wbook.close()